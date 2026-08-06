#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
YouTube 邮箱采集器 — 图形界面 (macOS / Windows 通用)
==================================================

把一个或多个 YouTube 频道链接粘贴到上方文本框，点击 “获取 / 加载”，
程序会抓取每个频道的：频道名称、频道链接、邮箱、订阅数。
完成后点击 “导出 Excel” 生成 .xlsx 表格。

依赖：requests, openpyxl  (Tkinter 随官方 Python 一并安装)
    pip install -r requirements.txt

运行：
    python youtube_email_gui.py
"""

from __future__ import annotations

import queue
import threading
import webbrowser

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog

import requests

# 复用命令行版本里的抓取逻辑
from youtube_email_scraper import (
    ChannelResult,
    EnrichmentConfig,
    about_url,
    extract_emails,
    scrape_channel,
    HAVE_ENRICHMENT,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAVE_OPENPYXL = True
except ImportError:  # pragma: no cover
    HAVE_OPENPYXL = False


APP_TITLE = "YouTube 邮箱采集器"
COLUMNS = [
    ("channel_name", "频道名称", 220),
    ("channel_url", "频道链接", 300),
    ("emails", "邮箱", 230),
    ("subscribers", "订阅数", 120),
    ("status", "状态", 90),
]
STATUS_TEXT = {
    "ok": "成功",
    "no_email": "无邮箱",
    "verification_required": "需登录验证",
    "error": "错误",
}


class ScraperGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("960x640")
        self.root.minsize(820, 520)

        self.results: list[ChannelResult] = []
        self.msg_queue: "queue.Queue" = queue.Queue()
        self.worker: threading.Thread | None = None
        self.cancel_flag = threading.Event()

        self._build_ui()
        self.root.after(120, self._drain_queue)

    # ------------------------------------------------------------------ UI
    def _build_ui(self) -> None:
        pad = {"padx": 10, "pady": 6}

        # ---- 顶部：输入区 ----
        top = ttk.LabelFrame(self.root, text="① 粘贴 YouTube 频道链接（每行一个）")
        top.pack(fill="x", **pad)

        self.input_text = tk.Text(top, height=6, wrap="none", undo=True)
        self.input_text.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=8)
        self.input_text.insert(
            "1.0",
            "https://www.youtube.com/@TechOnEarth\n",
        )
        yscroll = ttk.Scrollbar(top, orient="vertical", command=self.input_text.yview)
        yscroll.pack(side="right", fill="y", pady=8, padx=(0, 8))
        self.input_text.configure(yscrollcommand=yscroll.set)

        # ---- 中部：选项 + 按钮 ----
        bar = ttk.Frame(self.root)
        bar.pack(fill="x", padx=10)

        self.scan_videos = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            bar,
            text="About 页无邮箱时，扫描最近视频简介",
            variable=self.scan_videos,
        ).pack(side="left")

        ttk.Label(bar, text="视频数:").pack(side="left", padx=(10, 2))
        self.video_count = tk.IntVar(value=10)
        ttk.Spinbox(bar, from_=1, to=50, width=4, textvariable=self.video_count).pack(
            side="left"
        )

        # Enrichment option
        self.enable_enrichment = tk.BooleanVar(value=False)
        enrich_text = "启用增强搜索（社交媒体+Link-in-bio+网站）"
        if not HAVE_ENRICHMENT:
            enrich_text += " [未安装]"
        ttk.Checkbutton(
            bar,
            text=enrich_text,
            variable=self.enable_enrichment,
            state="normal" if HAVE_ENRICHMENT else "disabled",
        ).pack(side="left", padx=(15, 0))

        self.btn_fetch = ttk.Button(bar, text="② 获取 / 加载", command=self.on_fetch)
        self.btn_fetch.pack(side="right", padx=(6, 0))
        self.btn_cancel = ttk.Button(
            bar, text="停止", command=self.on_cancel, state="disabled"
        )
        self.btn_cancel.pack(side="right")
        self.btn_clear = ttk.Button(bar, text="清空结果", command=self.on_clear)
        self.btn_clear.pack(side="right", padx=(0, 6))

        # ---- 结果表 ----
        mid = ttk.LabelFrame(self.root, text="③ 抓取结果（双击可打开频道）")
        mid.pack(fill="both", expand=True, **pad)

        self.tree = ttk.Treeview(
            mid, columns=[c[0] for c in COLUMNS], show="headings", selectmode="extended"
        )
        for key, label, width in COLUMNS:
            self.tree.heading(key, text=label)
            anchor = "center" if key in ("subscribers", "status") else "w"
            self.tree.column(key, width=width, anchor=anchor, stretch=(key == "channel_url"))
        self.tree.tag_configure("no_email", foreground="#b06f00")
        self.tree.tag_configure("verification_required", foreground="#8e5b00")
        self.tree.tag_configure("error", foreground="#c0392b")

        vs = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        hs = ttk.Scrollbar(mid, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vs.grid(row=0, column=1, sticky="ns")
        hs.grid(row=1, column=0, sticky="ew")
        mid.rowconfigure(0, weight=1)
        mid.columnconfigure(0, weight=1)
        self.tree.bind("<Double-1>", self._open_selected)
        self.tree.bind("<<TreeviewSelect>>", self._on_selection_changed)

        # ---- 底部：进度 + 导出 ----
        bottom = ttk.Frame(self.root)
        bottom.pack(fill="x", **pad)

        self.progress = ttk.Progressbar(bottom, mode="determinate")
        self.progress.pack(side="left", fill="x", expand=True)

        self.btn_manual_email = ttk.Button(
            bottom, text="人工录入邮箱", command=self.on_manual_email, state="disabled"
        )
        self.btn_manual_email.pack(side="right", padx=(8, 0))

        self.btn_open_verify = ttk.Button(
            bottom, text="打开验证页", command=self.on_open_verification, state="disabled"
        )
        self.btn_open_verify.pack(side="right", padx=(8, 0))

        self.btn_export = ttk.Button(
            bottom, text="④ 导出 Excel", command=self.on_export, state="disabled"
        )
        self.btn_export.pack(side="right", padx=(8, 0))

        self.status_var = tk.StringVar(value="就绪")
        ttk.Label(self.root, textvariable=self.status_var, anchor="w",
                  relief="sunken").pack(fill="x", side="bottom")

    # ------------------------------------------------------------ actions
    def on_fetch(self) -> None:
        if self.worker and self.worker.is_alive():
            return
        raw = self.input_text.get("1.0", "end").strip()
        urls = [ln.split("#", 1)[0].strip() for ln in raw.splitlines()]
        urls = [u for u in urls if u]
        # 去重保序
        seen = set()
        deduped = []
        for u in urls:
            if u not in seen:
                seen.add(u)
                deduped.append(u)
        urls = deduped
        if not urls:
            messagebox.showwarning(APP_TITLE, "请先粘贴至少一个频道链接。")
            return

        self.on_clear()
        self.cancel_flag.clear()
        self.btn_fetch.config(state="disabled")
        self.btn_cancel.config(state="normal")
        self.btn_export.config(state="disabled")
        self.progress.config(maximum=len(urls), value=0)

        video_limit = self.video_count.get() if self.scan_videos.get() else 0

        # Setup enrichment config
        enrichment_config = None
        if self.enable_enrichment.get() and HAVE_ENRICHMENT:
            enrichment_config = EnrichmentConfig(
                enabled=True,
                social_media=True,
                biolink=True,
                website_deep=True,
                community_posts=True,
                max_website_depth=2,
                max_website_pages=10,
            )

        self.worker = threading.Thread(
            target=self._work, args=(urls, video_limit, enrichment_config), daemon=True
        )
        self.worker.start()

    def on_cancel(self) -> None:
        self.cancel_flag.set()
        self.status_var.set("正在停止…")

    def on_clear(self) -> None:
        self.results.clear()
        self.tree.delete(*self.tree.get_children())
        self.progress.config(value=0)
        self.btn_export.config(state="disabled")
        self.btn_open_verify.config(state="disabled")
        self.btn_manual_email.config(state="disabled")
        self.status_var.set("就绪")

    def on_export(self) -> None:
        if not self.results:
            messagebox.showinfo(APP_TITLE, "没有可导出的数据。")
            return
        if not HAVE_OPENPYXL:
            messagebox.showerror(
                APP_TITLE, "未安装 openpyxl，无法导出 Excel。\n请运行: pip install openpyxl"
            )
            return
        path = filedialog.asksaveasfilename(
            title="导出 Excel",
            defaultextension=".xlsx",
            initialfile="youtube_emails.xlsx",
            filetypes=[("Excel 工作簿", "*.xlsx")],
        )
        if not path:
            return
        try:
            self._write_xlsx(path)
        except Exception as e:  # pragma: no cover
            messagebox.showerror(APP_TITLE, f"导出失败：{e}")
            return
        if messagebox.askyesno(APP_TITLE, f"已导出：\n{path}\n\n是否现在打开？"):
            webbrowser.open(f"file://{path}")

    def on_open_verification(self) -> None:
        selected = self._selected_result()
        if selected is None:
            return
        _index, result = selected
        webbrowser.open(about_url(result.channel_url))
        self.status_var.set("已打开频道 About 页，请按 YouTube 正常流程登录并完成验证。")

    def on_manual_email(self) -> None:
        selected = self._selected_result()
        if selected is None:
            return
        index, result = selected
        value = simpledialog.askstring(
            APP_TITLE,
            "完成 YouTube 登录/验证后，把页面显示的邮箱粘贴到这里：",
            parent=self.root,
        )
        if value is None:
            return
        emails = extract_emails(value)
        if not emails:
            messagebox.showwarning(APP_TITLE, "没有识别到有效邮箱，请检查后重试。")
            return

        result.emails = emails
        result.source = "manual_after_verification"
        result.status = "ok"
        result.error = ""
        self._refresh_result_row(index, result)
        self.status_var.set(f"已人工录入：{result.emails_str}")
        self.btn_export.config(state="normal")

    # ------------------------------------------------------------- worker
    def _work(self, urls: list[str], video_limit: int, enrichment_config=None) -> None:
        session = requests.Session()
        for idx, url in enumerate(urls, 1):
            if self.cancel_flag.is_set():
                self.msg_queue.put(("done", "已停止"))
                return

            # Update status message
            status_msg = f"[{idx}/{len(urls)}] 抓取中：{url}"
            if enrichment_config and enrichment_config.enabled:
                status_msg += " (增强模式)"
            self.msg_queue.put(("status", status_msg))

            try:
                res = scrape_channel(
                    session, url,
                    video_limit=video_limit,
                    enrichment=enrichment_config,
                    cache=None  # No cache in GUI mode for now
                )
            except Exception as e:  # 兜底，避免线程崩溃
                res = ChannelResult(input=url, channel_url=url,
                                    status="error", error=str(e))
            self.msg_queue.put(("result", res))
        self.msg_queue.put(("done", None))

    def _drain_queue(self) -> None:
        try:
            while True:
                kind, payload = self.msg_queue.get_nowait()
                if kind == "status":
                    self.status_var.set(payload)
                elif kind == "result":
                    self._add_result(payload)
                elif kind == "done":
                    self._finish(payload)
        except queue.Empty:
            pass
        self.root.after(120, self._drain_queue)

    def _add_result(self, res: ChannelResult) -> None:
        self.results.append(res)
        index = len(self.results) - 1
        tag = res.status if res.status in (
            "no_email", "verification_required", "error"
        ) else ""
        self.tree.insert(
            "", "end", iid=str(index),
            values=self._result_values(res),
            tags=(tag,) if tag else (),
        )
        self.progress.step(1)

    def _finish(self, note: str | None) -> None:
        self.btn_fetch.config(state="normal")
        self.btn_cancel.config(state="disabled")
        if self.results:
            self.btn_export.config(state="normal")
        found = sum(1 for r in self.results if r.status == "ok")
        msg = note or f"完成：共 {len(self.results)} 个频道，找到邮箱 {found} 个。"
        self.status_var.set(msg)

    # ------------------------------------------------------------- export
    def _write_xlsx(self, path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "YouTube 邮箱"
        headers = ["YouTube频道名称", "YouTube频道链接", "YouTuber邮箱", "订阅数据", "状态"]
        ws.append(headers)

        head_font = Font(bold=True, color="FFFFFF")
        head_fill = PatternFill("solid", fgColor="C00000")
        for col, _ in enumerate(headers, 1):
            c = ws.cell(row=1, column=col)
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center", vertical="center")

        for r in self.results:
            ws.append([
                r.channel_name,
                r.channel_url,
                r.emails_str,
                r.subscribers,
                STATUS_TEXT.get(r.status, r.status),
            ])

        widths = [30, 46, 34, 18, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:E{ws.max_row}"
        wb.save(path)

    # ------------------------------------------------------------- helpers
    def _result_values(self, res: ChannelResult) -> tuple[str, ...]:
        empty_email = {
            "no_email": "—",
            "verification_required": "请打开验证页后人工录入",
        }.get(res.status, res.error)
        return (
            res.channel_name or "(未知)",
            res.channel_url,
            res.emails_str or empty_email,
            res.subscribers or "",
            STATUS_TEXT.get(res.status, res.status),
        )

    def _selected_result(self) -> tuple[int, ChannelResult] | None:
        selection = self.tree.selection()
        if not selection:
            return None
        try:
            index = int(selection[0])
            return index, self.results[index]
        except (ValueError, IndexError):
            return None

    def _refresh_result_row(self, index: int, res: ChannelResult) -> None:
        tag = res.status if res.status in (
            "no_email", "verification_required", "error"
        ) else ""
        self.tree.item(
            str(index), values=self._result_values(res), tags=(tag,) if tag else ()
        )

    def _on_selection_changed(self, _event=None) -> None:
        state = "normal" if self._selected_result() is not None else "disabled"
        self.btn_open_verify.config(state=state)
        self.btn_manual_email.config(state=state)

    def _open_selected(self, _event=None) -> None:
        selected = self._selected_result()
        if selected is None:
            return
        _index, result = selected
        url = (about_url(result.channel_url)
               if result.status == "verification_required" else result.channel_url)
        if url.startswith("http"):
            webbrowser.open(url)


def main() -> None:
    root = tk.Tk()
    try:
        # macOS 高分屏 / 通用主题
        style = ttk.Style()
        if "aqua" in style.theme_names():
            style.theme_use("aqua")
    except tk.TclError:
        pass
    ScraperGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
